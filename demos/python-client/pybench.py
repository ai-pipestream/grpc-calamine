# SPDX-License-Identifier: Apache-2.0
#
# NOT A DEMO. This is part of the benchmark harness in ../../bench and lives
# here only because it shares the generated stubs run.sh produces. It prints
# timings, not rows. To learn the API, read client.py instead; to run this,
# see ../../bench/README.md ("The other-language arms").
"""How does grpc-calamine compare to reading the same sheet in Python?

Three ways a Python program can get these cells:

  P1  openpyxl, read_only + values_only     the common answer
  P2  python-calamine                        the same Rust parser, bound in-process
  P3  grpc-calamine over gRPC                this project

Every arm counts rows and cells and accumulates the same order-sensitive
checksum over the values, so an arm that skips work cannot post a number.
Empty cells are normalised to None so the three agree on what a cell is.
"""

import os
import sys
import time
import zlib

CHECK = 0xCBF29CE484222325
MASK = (1 << 64) - 1


class Digest:
    __slots__ = ("h", "rows", "cells")

    def __init__(self):
        self.h = CHECK
        self.rows = 0
        self.cells = 0

    def _w(self, b):
        # zlib.crc32 runs in C. A pure-Python byte loop here would cost more
        # than the parser being measured and would make every arm look slow.
        self.h = zlib.crc32(b, self.h & 0xFFFFFFFF)

    def row(self):
        self.rows += 1

    def cell(self, v):
        self.cells += 1
        # Normalise to a stable byte form. Floats that are integral are written
        # as integers, because the three readers disagree on 72 vs 72.0.
        if v is None or v == "":
            self._w(b"\x00")
        elif isinstance(v, bool):
            self._w(b"\x05" + (b"\x01" if v else b"\x00"))
        elif isinstance(v, (int, float)):
            f = float(v)
            if f.is_integer():
                self._w(b"\x01" + int(f).to_bytes(8, "little", signed=True))
            else:
                import struct
                self._w(b"\x02" + struct.pack("<d", f))
        else:
            self._w(b"\x03" + str(v).encode("utf-8"))

    def __str__(self):
        return f"{self.h:016x}/{self.rows}r/{self.cells}c"


def p1_openpyxl(path, sheet):
    import openpyxl
    t = time.perf_counter()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    d = Digest()
    for row in ws.iter_rows(values_only=True):
        d.row()
        for v in row:
            d.cell(v)
    wb.close()
    return (time.perf_counter() - t) * 1e3, d


def p2_python_calamine(path, sheet):
    from python_calamine import CalamineWorkbook
    t = time.perf_counter()
    wb = CalamineWorkbook.from_path(path)
    ws = wb.get_sheet_by_name(sheet)
    d = Digest()
    for row in ws.to_python(skip_empty_area=False):
        d.row()
        for v in row:
            d.cell(v)
    return (time.perf_counter() - t) * 1e3, d


def p3_grpc(path, sheet, addr, use_dict=False):
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), "gen"))
    import grpc
    from calamine.v1 import calamine_service_pb2 as svc
    from calamine.v1 import calamine_service_pb2_grpc as rpc
    from calamine.v1 import types_pb2 as types

    CHUNK = 1 << 20
    data = open(path, "rb").read()

    def frames():
        yield svc.OpenWorkbookRequest(
            options=svc.WorkbookOptions(
                format_hint=types.WORKBOOK_FORMAT_UNSPECIFIED
            )
        )
        for i in range(0, len(data), CHUNK):
            yield svc.OpenWorkbookRequest(chunk=data[i : i + CHUNK])

    with grpc.insecure_channel(
        addr, options=[("grpc.max_receive_message_length", 32 << 20)]
    ) as ch:
        stub = rpc.CalamineServiceStub(ch)
        t_up = time.perf_counter()
        opened = stub.OpenWorkbook(frames())
        upload_ms = (time.perf_counter() - t_up) * 1e3

        t = time.perf_counter()
        d = Digest()
        # In dictionary mode ids resolve against an append-only list; the
        # contract guarantees every id is defined before its first use.
        table = []

        # Width of the last row taken, so an expanded gap is as wide as the
        # rows around it and the digest stays comparable with the other arms.
        width = [0]

        def take(r):
            d.row()
            width[0] = len(r.values)
            for c in r.values:
                which = c.WhichOneof("value")
                if which in ("string_value", "shared_string_value"):
                    d.cell(getattr(c, which))
                elif which == "shared_string_id":
                    d.cell(table[c.shared_string_id])
                elif which == "float_value":
                    d.cell(c.float_value)
                elif which == "int_value":
                    d.cell(c.int_value)
                elif which == "bool_value":
                    d.cell(c.bool_value)
                elif which == "date_time":
                    d.cell(c.date_time.value)
                elif which in ("date_time_iso", "duration_iso"):
                    d.cell(getattr(c, which))
                else:
                    d.cell(None)

        req = svc.StreamWorksheetRangeRequest(
            workbook_id=opened.workbook_id,
            sheet=svc.SheetSelector(sheet_name=sheet),
            max_rows_per_message=int(os.environ.get("BATCH", "0")),
            use_string_table=use_dict,
        )
        msgs = 0
        for ev in stub.StreamWorksheetRange(req):
            k = ev.WhichOneof("event")
            if k == "rows":
                msgs += 1
                for r in ev.rows.rows:
                    take(r)
            elif k == "row":
                msgs += 1
                take(ev.row)
            elif k == "row_gap":
                # The other arms densify, so this one has to as well or the
                # digests stop being comparable. That the expansion happens
                # here rather than on the wire is the saving being measured:
                # the gap is one message however many rows it stands for.
                msgs += 1
                for _ in range(ev.row_gap.row_count):
                    d.row()
                    for _ in range(width[0]):
                        d.cell(None)
            elif k == "string_table":
                assert ev.string_table.first_id == len(table), "chunks arrive in id order"
                table.extend(ev.string_table.entries)
        ms = (time.perf_counter() - t) * 1e3
        stub.CloseWorkbook(svc.CloseWorkbookRequest(workbook_id=opened.workbook_id))
        return ms, d, upload_ms, msgs


def main():
    path = sys.argv[1]
    sheet = sys.argv[2]
    addr = os.environ.get("CALAMINE_ADDR", "127.0.0.1:50055")
    only = os.environ.get("ONLY")

    print(f"workbook : {path}\nsheet    : {sheet}\n")
    results = []

    if only in (None, "p3"):
        ms, d, up, msgs = p3_grpc(path, sheet, addr)
        results.append(("P3 grpc-calamine over gRPC", ms, d))
        print(f"  upload once: {up:.0f} ms, {msgs} messages")

    if only in (None, "p4"):
        ms, d, up, msgs = p3_grpc(path, sheet, addr, use_dict=True)
        results.append(("P4 gRPC + use_string_table", ms, d))
        print(f"  upload once: {up:.0f} ms, {msgs} messages (dict)")

    if only in (None, "p2"):
        ms, d = p2_python_calamine(path, sheet)
        results.append(("P2 python-calamine (in-process)", ms, d))

    if only in (None, "p1"):
        ms, d = p1_openpyxl(path, sheet)
        results.append(("P1 openpyxl read_only", ms, d))

    print("\nsame-work proof")
    for name, _, d in results:
        print(f"  {name:34} {d}")
    hashes = {d.h for _, _, d in results}
    print(f"  identical: {'yes' if len(hashes) == 1 else 'NO'}\n")

    base = None
    print("wall clock")
    for name, ms, d in sorted(results, key=lambda r: r[1]):
        if base is None:
            base = ms
        print(
            f"  {name:34} {ms:9.0f} ms   {d.rows / (ms / 1000):9.0f} rows/s"
            f"   {ms / base:5.2f}x"
        )


if __name__ == "__main__":
    main()
